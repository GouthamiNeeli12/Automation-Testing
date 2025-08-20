from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time

fruit_name='Apple'
service=Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_argument('--start-maximized')
driver=webdriver.Chrome(service=service,options=options)

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.implicitly_wait(5)
driver.find_element(By.CSS_SELECTOR,"#downloadButton").click()

time.sleep(5)
#modify the excel file here
workbook=openpyxl.load_workbook("C:/Users/gouth/Downloads/download.xlsx")
sheet=workbook.active
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=1,column=i).value=='price':
        col=i

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        if sheet.cell(row=i,column=j).value=='Apple':
            row=i

sheet.cell(row=row,column=col).value=500
workbook.save("C:/Users/gouth/Downloads/download.xlsx")


path_toFile="C:/Users/gouth/Downloads/download.xlsx"
driver.find_element(By.CSS_SELECTOR,"input[type='file']").send_keys(path_toFile)

wait=WebDriverWait(driver,10)
wait.until(EC.visibility_of_element_located((By.XPATH,"//div[text()='Updated Excel Data Successfully.']")))

price_col_no=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price=driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_col_no+"-undefined']").text

print(actual_price)
driver.close()