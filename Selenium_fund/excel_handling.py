import openpyxl
workbook=openpyxl.load_workbook("C:/Users/gouth/Downloads/Book1.xlsx")
sheet=workbook.active
cell=sheet.cell(row=1,column=1)
print(cell.value)
sheet.cell(row=4,column=3).value="siri"
workbook.save("C:/Users/gouth/Downloads/Book1.xlsx")

print(sheet.max_row)
print(sheet.max_column)
print(sheet['B3'].value)
dict1={}
for row in range(1,sheet.max_row+1):
    # if sheet.cell(row=row,column=1).value=='Testcases':
    #     continue
    # else:
    for column in range(1,sheet.max_column+1):
            #print(sheet.cell(row=row,column=column).value,end=" ")
        dict1[sheet.cell(row=1,column=column).value]=sheet.cell(row=row,column=column).value

for key,value in dict1.items():
    print(key,value)