import openpyxl

workbook=openpyxl.load_workbook("C:/Users/gouth/Downloads/download.xlsx")
sheet=workbook.active
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=1,column=i).value=='price':
        col=i

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        if sheet.cell(row=i,column=j).value=='Apple':
            row=i

sheet.cell(row=row,column=col).value=500
workbook.save("C:/Users/gouth/Downloads/download.xlsx")